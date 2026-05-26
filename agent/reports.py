"""
Report generators for the business fleet quote.

generate_financial_excel(result) -> bytes   xlsx with Summary + per-role + Network sheets
generate_financial_pdf(result)   -> bytes   PDF BOM report
generate_rfp_pdf(result)         -> bytes   Vendor-facing RFP document
"""

from io import BytesIO
from datetime import datetime

# ── Excel ─────────────────────────────────────────────────────────────────────

def generate_financial_excel(result: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    BLUE  = "1F4E79"
    LBLUE = "D6E4F0"
    GREY  = "F2F2F2"
    WHITE = "FFFFFF"

    def _hdr(ws, row, col, text, bold=True, bg=BLUE, fg="FFFFFF", size=11):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font      = Font(bold=bold, color=fg, size=size)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        return cell

    def _val(ws, row, col, text, bold=False, bg=WHITE, num_format=None):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font      = Font(bold=bold, size=10)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(vertical="center")
        if num_format:
            cell.number_format = num_format
        return cell

    def _thin_border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def _apply_border(ws, min_row, max_row, min_col, max_col):
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(row=r, column=c).border = _thin_border()

    wb  = Workbook()
    cp  = result.get("company_profile", {})
    rrs = result.get("role_results", {})
    net = result.get("network") or {}
    now = datetime.now().strftime("%d %b %Y")

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    # Title block
    ws.merge_cells("A1:E1")
    _hdr(ws, 1, 1, "FLEET QUOTE — FINANCIAL SUMMARY", size=13)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:E2")
    c = ws.cell(row=2, column=1,
                value=f"{cp.get('name','?')}  ·  {cp.get('industry','?')}  ·  {cp.get('location','?')}  ·  Generated {now}")
    c.font      = Font(size=10, italic=True)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 8  # spacer

    # Column headers
    for ci, h in enumerate(["Role", "Headcount", "Per-Unit (RM)", "Role Total (RM)", "Notes"], 1):
        _hdr(ws, 4, ci, h)
    ws.row_dimensions[4].height = 22

    row = 5
    for role, data in rrs.items():
        count = data.get("count", 0)
        if "error" in data:
            for ci, v in enumerate([role.title(), count, "—", "—", data["error"]], 1):
                _val(ws, row, ci, v, bg=GREY)
        else:
            per_unit   = data.get("per_unit", 0)
            role_total = data.get("role_total", 0)
            notes      = ""
            warnings   = data.get("build", {}).get("warnings", [])
            if any("OVER BUDGET" in str(w) for w in warnings):
                notes = "⚠ Over budget"
            bg = "FFF2CC" if notes else WHITE
            for ci, v in enumerate([role.title(), count, per_unit, role_total, notes], 1):
                fmt = '#,##0.00' if ci in (3, 4) else None
                _val(ws, row, ci, v, bg=bg, num_format=fmt)
        row += 1

    # Totals
    net_total  = net.get("estimated_total_rm", 0)
    pc_total   = result.get("total_cost", 0)
    has_network = bool(net)

    ws.row_dimensions[row].height = 4  # spacer
    row += 1
    _hdr(ws, row, 1, "PC Fleet Total", bg=LBLUE, fg="000000")
    ws.merge_cells(f"B{row}:C{row}")
    t = ws.cell(row=row, column=4, value=pc_total)
    t.number_format = '#,##0.00'
    t.font = Font(bold=True, size=11)
    ws.cell(row=row, column=5, value=f"{result.get('total_pcs', 0)} PCs total").font = Font(italic=True, size=9)
    row += 1

    if has_network:
        _hdr(ws, row, 1, "Network Advisory (est.)", bg=LBLUE, fg="000000")
        ws.merge_cells(f"B{row}:C{row}")
        t = ws.cell(row=row, column=4, value=net_total)
        t.number_format = '#,##0.00'
        t.font = Font(bold=True, size=11)
        ws.cell(row=row, column=5, value="Separate budget line").font = Font(italic=True, size=9)
        row += 1

    _hdr(ws, row, 1, "Combined Estimate" if has_network else "Grand Total", bg=BLUE, fg="FFFFFF")
    ws.merge_cells(f"B{row}:C{row}")
    t = ws.cell(row=row, column=4, value=pc_total + net_total)
    t.number_format = '#,##0.00'
    t.font = Font(bold=True, size=12)
    t.fill = PatternFill("solid", fgColor="E2EFDA")

    _apply_border(ws, 4, row, 1, 5)

    # ── Per-role sheets ───────────────────────────────────────────────────────
    for role, data in rrs.items():
        title  = role.title()[:31]
        ws_r   = wb.create_sheet(title=title)
        for ci, w in zip("ABCDE", [28, 22, 18, 14, 14]):
            ws_r.column_dimensions[ci].width = w

        ws_r.merge_cells("A1:E1")
        _hdr(ws_r, 1, 1, f"{title.upper()} — BILL OF MATERIALS", size=12)
        ws_r.row_dimensions[1].height = 28

        if "error" in data:
            ws_r.cell(row=2, column=1, value=f"Error: {data['error']}")
            continue

        count    = data.get("count", 0)
        per_unit = data.get("per_unit", 0)
        build    = data.get("build", {})
        costs    = build.get("costs", {})

        ws_r.cell(row=2, column=1,
                  value=f"Headcount: {count}  ·  Per-unit total: RM {per_unit:,.2f}  ·  "
                        f"Role total: RM {data.get('role_total', 0):,.2f}").font = Font(size=9, italic=True)

        for ci, h in enumerate(["Category", "Component", "Vendor", "Unit Price (RM)", f"×{count} Total (RM)"], 1):
            _hdr(ws_r, 4, ci, h)
        ws_r.row_dimensions[4].height = 20

        row = 5
        for item in build.get("items", []):
            cat      = item.get("category", "").replace("_", " ").title()
            name     = item.get("name", "—")
            vendor   = item.get("vendor_name", "—")
            up       = item.get("unit_price_rm", 0)
            row_tot  = round(up * count, 2)
            bg = GREY if row % 2 == 0 else WHITE
            for ci, v in enumerate([cat, name, vendor, up, row_tot], 1):
                fmt = '#,##0.00' if ci in (4, 5) else None
                _val(ws_r, row, ci, v, bg=bg, num_format=fmt)
            row += 1

        row += 1  # spacer
        cost_rows = [
            ("Subtotal",     costs.get("subtotal_rm", 0)),
            ("Shipping",     costs.get("shipping_total_rm", 0)),
            ("SST (6%)",     costs.get("sst_rm", 0)),
            ("Grand Total",  costs.get("grand_total_rm", 0)),
        ]
        for label, val in cost_rows:
            bold = label == "Grand Total"
            bg   = LBLUE if label == "Grand Total" else WHITE
            ws_r.merge_cells(f"A{row}:C{row}")
            _val(ws_r, row, 1, label, bold=bold, bg=bg)
            t = ws_r.cell(row=row, column=4, value=val)
            t.number_format = '#,##0.00'
            t.font = Font(bold=bold, size=10)
            t.fill = PatternFill("solid", fgColor=bg)
            t2 = ws_r.cell(row=row, column=5, value=round(val * count, 2))
            t2.number_format = '#,##0.00'
            t2.font = Font(bold=bold, size=10)
            t2.fill = PatternFill("solid", fgColor=bg)
            row += 1

        _apply_border(ws_r, 4, row - 1, 1, 5)

    # ── Network sheet ─────────────────────────────────────────────────────────
    ws_n = wb.create_sheet(title="Network")
    for ci, w in zip("ABCD", [20, 36, 16, 16]):
        ws_n.column_dimensions[ci].width = w

    ws_n.merge_cells("A1:D1")
    _hdr(ws_n, 1, 1, "NETWORK INFRASTRUCTURE — ADVISORY BOM", size=12)
    ws_n.row_dimensions[1].height = 28

    ws_n.cell(row=2, column=1,
              value="Advisory only — separate budget line from PC fleet.").font = Font(size=9, italic=True)

    for ci, h in enumerate(["Item", "Details", "Qty", "Price (RM)"], 1):
        _hdr(ws_n, 4, ci, h)

    net_rows = []
    sw = net.get("switch")
    if sw:
        net_rows.append(("Switch", f"{sw['name']} — {sw.get('description','')}",
                         sw.get("quantity", 1), sw.get("subtotal_rm", 0)))
    rt = net.get("router")
    if rt:
        net_rows.append(("Router", f"{rt['name']} — {rt.get('description','')}",
                         1, rt.get("price_rm", 0)))
    nas = net.get("nas")
    if nas:
        net_rows.append(("NAS", f"{nas['name']} — {nas.get('description','')}",
                         1, nas.get("price_rm", 0)))
    wifi = net.get("wifi")
    if wifi:
        net_rows.append(("WiFi APs",
                         f"{wifi.get('recommendation','')} @ RM{wifi.get('unit_price_rm',0):,}/unit",
                         wifi.get("access_points_qty", 0), wifi.get("subtotal_rm", 0)))
    cabling = net.get("cabling")
    if cabling:
        net_rows.append(("Cat6 Cabling",
                         f"~{cabling.get('estimated_metres',0)} m · "
                         f"{cabling.get('bulk_boxes_qty',0)} box(es) + "
                         f"{cabling.get('patch_cables_qty',0)} patch cables",
                         1, cabling.get("subtotal_rm", 0)))

    row = 5
    for item_name, details, qty, price in net_rows:
        bg = GREY if row % 2 == 0 else WHITE
        for ci, v in enumerate([item_name, details, qty, price], 1):
            fmt = '#,##0.00' if ci == 4 else None
            _val(ws_n, row, ci, v, bg=bg, num_format=fmt)
        row += 1

    row += 1
    ws_n.merge_cells(f"A{row}:C{row}")
    _val(ws_n, row, 1, "Network Total", bold=True, bg=LBLUE)
    t = ws_n.cell(row=row, column=4, value=net.get("estimated_total_rm", 0))
    t.number_format = '#,##0.00'
    t.font = Font(bold=True, size=11)
    t.fill = PatternFill("solid", fgColor=LBLUE)

    _apply_border(ws_n, 4, row, 1, 4)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF helpers ────────────────────────────────────────────────────────────────

def _pdf_styles():
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle("DocTitle",
        parent=styles["Title"], fontSize=20, spaceAfter=6, textColor=colors.HexColor("#1F4E79")))
    styles.add(ParagraphStyle("SectionTitle",
        parent=styles["Heading1"], fontSize=13, spaceBefore=14, spaceAfter=4,
        textColor=colors.HexColor("#1F4E79"),
        borderPad=4, borderWidth=0, borderColor=colors.HexColor("#1F4E79")))
    styles.add(ParagraphStyle("SubTitle",
        parent=styles["Heading2"], fontSize=11, spaceBefore=8, spaceAfter=2,
        textColor=colors.HexColor("#2E75B6")))
    styles.add(ParagraphStyle("Body",
        parent=styles["Normal"], fontSize=9, spaceAfter=4, leading=13))
    styles.add(ParagraphStyle("SmallItalic",
        parent=styles["Normal"], fontSize=8, italic=True,
        textColor=colors.HexColor("#555555")))
    styles.add(ParagraphStyle("TableHeader",
        parent=styles["Normal"], fontSize=9, bold=True, textColor=colors.white))
    return styles


def _pdf_table(data, col_widths, hdr_bg="#1F4E79"):
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors

    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor(hdr_bg)),
        ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0),  9),
        ("FONTNAME",    (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",    (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF3FB")]),
        ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
    ]))
    return tbl


# ── Financial PDF ──────────────────────────────────────────────────────────────

def generate_financial_pdf(result: dict) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, HRFlowable
    from reportlab.lib.units import cm
    from reportlab.lib import colors

    buf    = BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=A4,
                               leftMargin=2*cm, rightMargin=2*cm,
                               topMargin=2.5*cm, bottomMargin=2*cm)
    styles = _pdf_styles()
    W      = A4[0] - 4*cm  # usable width
    cp     = result.get("company_profile", {})
    rrs    = result.get("role_results", {})
    net    = result.get("network", {})
    now    = datetime.now().strftime("%d %B %Y")
    elems  = []

    # Cover header
    elems.append(Paragraph("Fleet Quote — Financial Report", styles["DocTitle"]))
    elems.append(Paragraph(
        f"<b>{cp.get('name','?')}</b>  ·  {cp.get('industry','?')}  ·  "
        f"{cp.get('location','?')}  ·  Generated {now}",
        styles["SmallItalic"]
    ))
    elems.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#1F4E79"), spaceAfter=10))

    # ── 1. Fleet Summary ──────────────────────────────────────────────────────
    elems.append(Paragraph("1. Fleet Summary", styles["SectionTitle"]))
    summary_data = [["Role", "Headcount", "Per Unit (RM)", "Role Total (RM)"]]
    for role, data in rrs.items():
        if "error" in data:
            summary_data.append([role.title(), data.get("count", 0), "Error", "—"])
        else:
            summary_data.append([
                role.title(),
                str(data.get("count", 0)),
                f"RM {data.get('per_unit', 0):,.2f}",
                f"RM {data.get('role_total', 0):,.2f}",
            ])
    elems.append(_pdf_table(summary_data, [W*0.3, W*0.15, W*0.25, W*0.3]))
    elems.append(Spacer(1, 6))

    net_total = net.get("estimated_total_rm", 0)
    pc_cost   = result.get("total_cost", 0)
    if net:
        totals_data = [
            ["", "PC Fleet Total", f"RM {pc_cost:,.2f}"],
            ["", "Network Advisory (separate)", f"RM {net_total:,.2f}"],
            ["", "Combined Estimate (informational)", f"RM {pc_cost + net_total:,.2f}"],
        ]
    else:
        totals_data = [
            ["", "PC Fleet Total", f"RM {pc_cost:,.2f}"],
        ]
    tbl2 = Table(totals_data, colWidths=[W*0.05, W*0.6, W*0.35])
    tbl2.setStyle(TableStyle([
        ("FONTNAME",   (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("FONTNAME",   (1, 2), (2, 2),   "Helvetica-Bold"),
        ("FONTSIZE",   (1, 2), (2, 2),   10),
        ("BACKGROUND", (0, 2), (-1, 2),  colors.HexColor("#D6E4F0")),
        ("ALIGN",      (2, 0), (2, -1),  "RIGHT"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elems.append(tbl2)

    # ── 2. Per-role BOMs ──────────────────────────────────────────────────────
    elems.append(Paragraph("2. Per-Role Bills of Materials", styles["SectionTitle"]))

    for role, data in rrs.items():
        elems.append(Paragraph(f"{role.title()}", styles["SubTitle"]))
        if "error" in data:
            elems.append(Paragraph(f"Generation failed: {data['error']}", styles["Body"]))
            continue

        count  = data.get("count", 0)
        build  = data.get("build", {})
        costs  = build.get("costs", {})
        elems.append(Paragraph(
            f"Headcount: {count}  ·  Per unit: RM {data.get('per_unit',0):,.2f}  ·  "
            f"Role total: RM {data.get('role_total',0):,.2f}",
            styles["SmallItalic"]
        ))

        bom_data = [["Category", "Component", "Vendor", "Unit Price"]]
        for item in build.get("items", []):
            bom_data.append([
                item.get("category", "").replace("_", " ").title(),
                item.get("name", "—"),
                item.get("vendor_name", "—"),
                f"RM {item.get('unit_price_rm', 0):,.2f}",
            ])
        elems.append(_pdf_table(bom_data, [W*0.18, W*0.42, W*0.22, W*0.18]))
        elems.append(Spacer(1, 4))

        cost_tbl = [
            ["", "Subtotal",    f"RM {costs.get('subtotal_rm',0):,.2f}"],
            ["", "Shipping",    f"RM {costs.get('shipping_total_rm',0):,.2f}"],
            ["", "SST (6%)",    f"RM {costs.get('sst_rm',0):,.2f}"],
            ["", "Grand Total", f"RM {costs.get('grand_total_rm',0):,.2f}"],
        ]
        ct = Table(cost_tbl, colWidths=[W*0.6, W*0.2, W*0.2])
        ct.setStyle(TableStyle([
            ("FONTNAME",   (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE",   (0, 0), (-1, -1), 8),
            ("FONTNAME",   (1, 3), (2, 3),   "Helvetica-Bold"),
            ("BACKGROUND", (0, 3), (-1, 3),  colors.HexColor("#D6E4F0")),
            ("ALIGN",      (2, 0), (2, -1),  "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elems.append(ct)
        elems.append(Spacer(1, 8))

    # ── 3. Network ────────────────────────────────────────────────────────────
    if net:
        elems.append(Paragraph("3. Network Infrastructure", styles["SectionTitle"]))
        elems.append(Paragraph(
            net.get("disclaimer", "Advisory only — separate budget line."),
            styles["SmallItalic"]
        ))
        net_rows = [["Item", "Details", "Price (RM)"]]
        for key, label in [("switch","Switch"),("router","Router"),("nas","NAS")]:
            item = net.get(key)
            if item:
                price = item.get("subtotal_rm") or item.get("price_rm", 0)
                net_rows.append([label, f"{item['name']} — {item.get('description','')}", f"RM {price:,.0f}"])
        wifi = net.get("wifi")
        if wifi:
            net_rows.append(["WiFi APs",
                f"{wifi.get('recommendation','')} ×{wifi.get('access_points_qty',0)}",
                f"RM {wifi.get('subtotal_rm',0):,.0f}"])
        cabling = net.get("cabling")
        if cabling:
            net_rows.append(["Cat6 Cabling",
                f"~{cabling.get('estimated_metres',0)} m",
                f"RM {cabling.get('subtotal_rm',0):,.0f}"])
        net_rows.append(["", "Network Total", f"RM {net.get('estimated_total_rm',0):,.0f}"])

        nt = _pdf_table(net_rows, [W*0.15, W*0.6, W*0.25])
        nt.setStyle(TableStyle([
            ("FONTNAME",   (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#D6E4F0")),
        ]))
        elems.append(nt)

    doc.build(elems)
    return buf.getvalue()


# ── RFP PDF ────────────────────────────────────────────────────────────────────

def generate_rfp_pdf(result: dict) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle, PageBreak, HRFlowable)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors

    buf    = BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=A4,
                               leftMargin=2.5*cm, rightMargin=2.5*cm,
                               topMargin=3*cm, bottomMargin=2.5*cm)
    styles = _pdf_styles()
    W      = A4[0] - 5*cm
    cp     = result.get("company_profile", {})
    rrs    = result.get("role_results", {})
    net    = result.get("network") or {}
    now    = datetime.now().strftime("%d %B %Y")
    elems  = []

    total_pcs  = result.get("total_pcs", 0)
    total_cost = result.get("total_cost", 0)
    net_total  = net.get("estimated_total_rm", 0)

    # ── Cover Page ────────────────────────────────────────────────────────────
    elems.append(Spacer(1, 2*cm))
    elems.append(Paragraph("REQUEST FOR PROPOSAL", styles["DocTitle"]))
    elems.append(Paragraph("PC Fleet Procurement", ParagraphStyle(
        "Cover2", parent=styles["Normal"], fontSize=14, spaceAfter=6,
        textColor=colors.HexColor("#2E75B6"))))
    elems.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#1F4E79"), spaceAfter=20))
    cover_data = [
        ["Issued by:",    cp.get("name", "?")],
        ["Industry:",     cp.get("industry", "?")],
        ["Location:",     cp.get("location", "?")],
        ["Fleet size:",   f"{total_pcs} PCs across {len(rrs)} role(s)"],
        ["Budget (PCs):", f"RM {total_cost:,.0f}"],
        ["Issue date:",   now],
    ]
    ct = Table(cover_data, colWidths=[W*0.3, W*0.7])
    ct.setStyle(TableStyle([
        ("FONTNAME",    (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1,-1), 10),
        ("TOPPADDING",  (0, 0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1), 5),
        ("ROWBACKGROUNDS", (0,0),(-1,-1),[colors.white, colors.HexColor("#F5F9FF")]),
    ]))
    elems.append(ct)
    elems.append(Spacer(1, 1*cm))
    elems.append(Paragraph(
        "This document is a formal Request for Proposal issued to prospective vendors "
        "for the supply and delivery of PC hardware for the above company's fleet. "
        "Vendors are invited to submit competitive quotations addressing all sections below.",
        styles["Body"]
    ))
    elems.append(PageBreak())

    # ── 1. Executive Summary ──────────────────────────────────────────────────
    elems.append(Paragraph("1. Executive Summary", styles["SectionTitle"]))
    elems.append(Paragraph(
        f"<b>{cp.get('name','?')}</b> is a {cp.get('size','?')}-person "
        f"{cp.get('industry','?')} company based in {cp.get('location','?')}. "
        f"The company is procuring a PC fleet of <b>{total_pcs} units</b> across "
        f"<b>{len(rrs)} distinct roles</b> to support its operations.",
        styles["Body"]
    ))
    elems.append(Paragraph(
        f"The estimated PC fleet budget is <b>RM {total_cost:,.0f}</b>. "
        f"An additional network infrastructure budget of <b>RM {net_total:,.0f}</b> "
        "(advisory, separate procurement) is also provided for reference.",
        styles["Body"]
    ))
    elems.append(Paragraph(
        "Procurement objectives: (1) Acquire hardware matched to each role's workload; "
        "(2) Stay within the per-role budget targets; "
        "(3) Ensure compatibility and warranty coverage; "
        "(4) Complete delivery within the agreed timeline.",
        styles["Body"]
    ))

    # ── 2. Scope of Supply ────────────────────────────────────────────────────
    elems.append(Paragraph("2. Scope of Supply", styles["SectionTitle"]))
    elems.append(Paragraph(
        "Vendors shall supply all hardware components as itemised per role in Section 4. "
        "Supply must include packaging, warranty registration, and delivery to the address below.",
        styles["Body"]
    ))
    scope_data = [["Role", "Headcount", "Target Budget/Unit", "Key Workload"]]
    for role, data in rrs.items():
        count  = data.get("count", 0)
        per_u  = f"RM {data.get('per_unit',0):,.0f}" if "error" not in data else "—"
        build  = data.get("build", {})
        rationale = build.get("build_rationale", "")[:80] + "…" if build.get("build_rationale","") else "—"
        scope_data.append([role.title(), str(count), per_u, rationale])
    elems.append(_pdf_table(scope_data, [W*0.18, W*0.12, W*0.2, W*0.5]))
    elems.append(Paragraph(
        f"Delivery location: {cp.get('location','?')}, Malaysia.",
        styles["SmallItalic"]
    ))

    # ── 3. Technical Specs ────────────────────────────────────────────────────
    elems.append(Paragraph("3. Technical Specifications per Role", styles["SectionTitle"]))
    elems.append(Paragraph(
        "The following specifications represent the minimum requirements per role. "
        "Vendors may propose equivalent or superior specifications with justification.",
        styles["Body"]
    ))
    for role, data in rrs.items():
        elems.append(Paragraph(f"3.{list(rrs.keys()).index(role)+1}  {role.title()}", styles["SubTitle"]))
        if "error" in data:
            elems.append(Paragraph(f"Specification unavailable: {data['error']}", styles["Body"]))
            continue
        build = data.get("build", {})
        elems.append(Paragraph(build.get("build_rationale", "No rationale available."), styles["Body"]))
        spec_data = [["Category", "Required Component / Spec", "Min Price Target"]]
        for item in build.get("items", []):
            spec_data.append([
                item.get("category","").replace("_"," ").title(),
                item.get("name","—"),
                f"RM {item.get('unit_price_rm',0):,.0f}",
            ])
        elems.append(_pdf_table(spec_data, [W*0.2, W*0.55, W*0.25]))
        elems.append(Spacer(1, 6))

    # ── 4. Full BOM ───────────────────────────────────────────────────────────
    elems.append(Paragraph("4. Full Bill of Materials", styles["SectionTitle"]))
    elems.append(Paragraph(
        "Complete consolidated BOM for all roles. Unit prices are reference targets "
        "based on current market data. Vendors should quote per-unit and per-fleet prices.",
        styles["Body"]
    ))
    bom_data = [["Role", "Category", "Component", "Vendor Ref.", "Qty", "Unit Price"]]
    for role, data in rrs.items():
        if "error" in data:
            continue
        count = data.get("count", 0)
        for item in data.get("build", {}).get("items", []):
            bom_data.append([
                role.title(),
                item.get("category","").replace("_"," ").title(),
                item.get("name","—"),
                item.get("vendor_name","—"),
                str(count),
                f"RM {item.get('unit_price_rm',0):,.0f}",
            ])
    elems.append(_pdf_table(bom_data, [W*0.12, W*0.14, W*0.34, W*0.16, W*0.08, W*0.16]))

    # ── 5. Delivery Requirements ──────────────────────────────────────────────
    elems.append(Paragraph("5. Delivery & Installation Requirements", styles["SectionTitle"]))
    delivery_items = [
        "All units must be delivered assembled and tested (POST passed) unless otherwise agreed.",
        "Original retail packaging required. Components must be individually boxed with serial numbers visible.",
        f"Delivery destination: {cp.get('name','?')}, {cp.get('location','?')}, Malaysia.",
        "Vendor to provide a packing list and delivery note for each shipment.",
        "Partial deliveries are acceptable provided roles are delivered complete (not split across shipments).",
        f"Suggested delivery timeline: within 14–21 business days of purchase order confirmation.",
        "Vendor responsible for safe transport; damaged goods must be replaced at no additional cost.",
    ]
    for item in delivery_items:
        elems.append(Paragraph(f"• {item}", styles["Body"]))
    elems.append(Spacer(1, 6))

    # ── 6. Pricing & Payment Terms ────────────────────────────────────────────
    elems.append(Paragraph("6. Pricing & Payment Terms", styles["SectionTitle"]))
    pricing_items = [
        "All prices must be quoted in Malaysian Ringgit (MYR / RM).",
        "Quotes must include SST (6%) and itemise shipping/handling charges separately.",
        "Prices valid for a minimum of 30 days from submission date.",
        "Suggested payment terms: 30% deposit upon PO confirmation; 70% upon delivery and acceptance.",
        "Invoices must reference the PO number and include itemised line items matching this BOM.",
        "Vendor may propose volume discounts for quantities ≥ 10 units per role.",
    ]
    for item in pricing_items:
        elems.append(Paragraph(f"• {item}", styles["Body"]))
    elems.append(Spacer(1, 6))

    # ── 7. Vendor Qualification ───────────────────────────────────────────────
    elems.append(Paragraph("7. Vendor Qualification Criteria", styles["SectionTitle"]))
    elems.append(Paragraph(
        "Submissions will be evaluated on the following criteria:",
        styles["Body"]
    ))
    qual_data = [
        ["Criterion", "Requirement"],
        ["Local presence",        "Registered business entity in Malaysia (SSM required)"],
        ["Authorised reseller",   "Must be authorised reseller or distributor for major brands quoted"],
        ["Warranty coverage",     "Minimum 1-year on-site or RTB warranty on all components"],
        ["After-sales support",   "Local support contact; response within 2 business days"],
        ["Track record",          "At least 2 verifiable B2B supply references in the past 2 years"],
        ["Component authenticity","All parts must be original, not refurbished, unless explicitly stated"],
        ["Financial stability",   "Ability to fulfil full order; bank reference or credit line if requested"],
    ]
    elems.append(_pdf_table(qual_data, [W*0.3, W*0.7], hdr_bg="#2E75B6"))

    # ── 8. Submission Instructions ────────────────────────────────────────────
    elems.append(Paragraph("8. Submission Instructions", styles["SectionTitle"]))
    sub_items = [
        "Quotation must address all roles and line items in Section 4 (partial quotes will not be considered).",
        "Submit as a single PDF or Excel file with company letterhead and authorised signature.",
        "Include company profile, SSM registration copy, and at least one B2B reference.",
        f"Send quotation to: [procurement contact — to be filled by {cp.get('name','?')}]",
        "Submission deadline: [to be specified by issuer]",
        "Late submissions may be considered at the issuer's discretion.",
        "The issuer reserves the right to accept or reject any quotation without explanation.",
    ]
    for item in sub_items:
        elems.append(Paragraph(f"• {item}", styles["Body"]))

    elems.append(Spacer(1, 1*cm))
    elems.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#AAAAAA")))
    elems.append(Paragraph(
        f"Document generated by AI PC Procurement Agent  ·  {now}  ·  "
        f"Session {result.get('session_id','')}",
        styles["SmallItalic"]
    ))

    doc.build(elems)
    return buf.getvalue()
